"""Excel exports of the manager lists (drivers, vehicles, companies).

Every cell is written as text on purpose: Excel would otherwise eat the leading
zero of an SSN and re-format dates to the reader's locale — the managers are
split between Windows and macOS, so the sheet must look the same on both.
"""

from datetime import date
from io import BytesIO
from typing import Annotated, Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlmodel import Session, select

import uploads
from database import get_session
from dependencies import get_current_user
from models import (
    Company,
    ComplianceDocument,
    Driver,
    DriverAnswers,
    DriverApplication,
    Truck,
    User,
)
from routers.compliance import current_docs

router = APIRouter(prefix="/api/export", tags=["Export"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def us(value: Optional[Any]) -> str:
    """Dates as MM/DD/YYYY; accepts the ISO strings stored in application answers."""
    if not value:
        return ""
    if isinstance(value, str):
        try:
            value = date.fromisoformat(value[:10])
        except ValueError:
            return value
    return value.strftime("%m/%d/%Y")


def sheet(name: str, headers: List[str], rows: List[List[str]]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers):
        width = max([len(header)] + [len(r[i]) for r in rows]) + 2
        ws.column_dimensions[get_column_letter(i + 1)].width = min(width, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'},
    )


def company_names(session: Session) -> Dict[int, str]:
    return {c.id: c.name for c in session.exec(select(Company)).all()}


def pick(rows: List[Any], wanted: Optional[List[int]], order_by) -> List[Any]:
    """Narrow a table to the rows the list page is showing, in the order it shows them.
    `wanted` is None when the page sent no selection — then everything, sorted by
    `order_by`. An empty list means the page is showing nothing, and so is the sheet.

    ponytail: the whole (small) table is read and narrowed in Python. That is what
    lets one endpoint honour every filter, search and sort the page has without
    restating any of them; if these tables ever reach five figures, filter in SQL.
    """
    if wanted is None:
        return sorted(rows, key=order_by)
    by_id = {r.id: r for r in rows}
    return [by_id[i] for i in wanted if i in by_id]


def export_drivers(session: Session, wanted: Optional[List[int]]) -> StreamingResponse:
    drivers = pick(
        list(session.exec(select(Driver)).all()), wanted,
        lambda d: (d.last_name.lower(), d.first_name.lower()),
    )
    ids = [d.id for d in drivers]
    names = company_names(session)

    # The licence copy on file is the current truth for state, number and expiry...
    cdl: Dict[int, ComplianceDocument] = {}
    if ids:
        for doc in session.exec(
            current_docs().where(
                ComplianceDocument.document_type == uploads.DOC_TYPES["cdl"],
                ComplianceDocument.driver_id.in_(ids),
            )
        ).all():
            cdl[doc.driver_id] = doc

    # ...falling back to the driver's application for anything not entered there.
    answers: Dict[int, dict] = {}
    if ids:
        rows = session.exec(
            select(DriverApplication.driver_id, DriverAnswers)
            .join(DriverAnswers, DriverAnswers.application_id == DriverApplication.id)
            .where(DriverApplication.driver_id.in_(ids))
            .order_by(DriverApplication.id)
        ).all()
        for driver_id, row in rows:
            answers[driver_id] = row.answers or {}   # later application wins

    out = []
    for d in drivers:
        doc = cdl.get(d.id)
        app_cdl = (answers.get(d.id) or {}).get("cdl") or {}
        out.append([
            d.first_name,
            d.middle_name or "",
            d.last_name,
            us(d.dob),
            (doc.issuing_state if doc and doc.issuing_state else app_cdl.get("state", "")) or "",
            (doc.document_number if doc and doc.document_number else app_cdl.get("number", "")) or "",
            us(doc.expiry_date if doc else app_cdl.get("expiration")),
            d.ssn or "",
            us(d.hire_date),
            names.get(d.company_id, ""),
        ])
    return sheet(
        "Drivers",
        ["First name", "Middle name", "Last name", "DOB", "License state",
         "License no", "Expire date", "Social", "Hire date", "Company"],
        out,
    )


def export_trucks(session: Session, wanted: Optional[List[int]]) -> StreamingResponse:
    trucks = pick(
        list(session.exec(select(Truck)).all()), wanted,
        lambda t: (t.unit_number or "", t.make),
    )
    names = company_names(session)
    rows = [
        [
            t.unit_number or "",
            t.make,
            str(t.year),
            t.vin,
            t.plate_number,
            t.state_registered,
            names.get(t.company_id, ""),
        ]
        for t in trucks
    ]
    return sheet(
        "Vehicles",
        ["Unit number", "Make", "Year", "VIN", "Plate number", "State", "Company"],
        rows,
    )


def export_companies(session: Session, wanted: Optional[List[int]]) -> StreamingResponse:
    companies = pick(
        list(session.exec(select(Company)).all()), wanted, lambda c: c.name.lower(),
    )
    rows = [
        [
            c.name,
            c.dot_number or "",
            c.mc_number or "",
            f"{c.address_street}, {c.address_city}, {c.address_state} {c.address_zip}".strip(),
            c.phone or "",
            c.owner_name or "",
            us(c.owner_dob),
        ]
        for c in companies
    ]
    return sheet(
        "Companies",
        ["Name", "DOT", "MC", "Location", "Phone", "Owner", "Owner DOB"],
        rows,
    )


class ExportRequest(BaseModel):
    """The rows the list page is showing, in its own order. Sending them spares the
    export from restating the page's filters (company, checklist, documents, status,
    search box) — whatever the manager sees is what lands in the sheet. Omit for all."""
    ids: Optional[List[int]] = None


@router.post("/{kind}")
def export(
    kind: str,
    session: Annotated[Session, Depends(get_session)],
    _user: Annotated[User, Depends(get_current_user)],
    body: ExportRequest = ExportRequest(),
) -> StreamingResponse:
    """`kind` is drivers | trucks | companies."""
    if kind == "drivers":
        return export_drivers(session, body.ids)
    if kind == "trucks":
        return export_trucks(session, body.ids)
    if kind == "companies":
        return export_companies(session, body.ids)
    raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Unknown export")
