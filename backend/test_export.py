"""Excel exports: right columns, US dates, SSN kept as text.
Run: python test_export.py   (needs DATABASE_URL; cleans up after itself)
"""

from io import BytesIO

from dotenv import load_dotenv

load_dotenv()

from PIL import Image  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, select  # noqa: E402

import cascade  # noqa: E402
from database import get_engine  # noqa: E402
from dependencies import get_current_user, get_current_user_file  # noqa: E402
from main import app  # noqa: E402
from models import Company, Driver, Truck, User  # noqa: E402

CO = "Export Test Co"


def _photo():
    buf = BytesIO()
    Image.new("RGB", (60, 40), "red").save(buf, "JPEG")
    return buf.getvalue()


def _clean(s):
    for co in s.exec(select(Company).where(Company.name == CO)).all():
        for d in s.exec(select(Driver).where(Driver.company_id == co.id)).all():
            cascade.delete_driver(s, d)
        for t in s.exec(select(Truck).where(Truck.company_id == co.id)).all():
            cascade.delete_truck(s, t)
        s.delete(co)
    s.commit()


def rows(client, path, ids=None):
    r = client.post(path, json={"ids": ids})
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"], r.headers
    ws = load_workbook(BytesIO(r.content)).active
    return [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]


def main():
    engine = get_engine()
    with Session(engine) as s:
        _clean(s)
        co = Company(name=CO, address_street="1 Main St", address_city="Dallas",
                     address_state="TX", address_zip="75001", phone="(214) 555-0100",
                     dot_number="1234567", mc_number="MC-9999",
                     owner_name="Jane Owner", owner_dob="1980-02-03")
        s.add(co); s.commit(); s.refresh(co)
        company_id = co.id

    fake = User(id=1, email="t@t.t", hashed_password="x", role="admin", is_active=True)
    app.dependency_overrides[get_current_user] = lambda: fake
    app.dependency_overrides[get_current_user_file] = lambda: fake
    client = TestClient(app)

    try:
        did = client.post("/api/drivers", json={
            "company_id": company_id, "first_name": "Bob", "middle_name": "Q",
            "last_name": "Rae", "email": "b@r.com", "phone": "2145550101",
            "dob": "1990-01-02", "hire_date": "2024-03-04", "ssn": "012-34-5678",
        }).json()["id"]
        # The manager can fix both afterwards; the list DTO still must not carry an SSN.
        assert client.patch(f"/api/drivers/{did}", json={"dob": "1990-01-03"}).json()["dob"] == "1990-01-03"
        assert client.get(f"/api/drivers/{did}").json()["ssn"] == "012-34-5678"
        assert "ssn" not in client.get("/api/drivers").json()[0]
        client.post(f"/api/drivers/{did}/documents/cdl",
                    data={"expiry": "2030-12-31", "number": "TX-777", "issuing_state": "tx"},
                    files={"file": ("dl.jpg", _photo(), "image/jpeg")})
        # A later correction must not blank the state entered with the file.
        client.post(f"/api/drivers/{did}/documents/cdl", data={"expiry": "2031-01-31"})
        client.post("/api/trucks", json={
            "company_id": company_id, "make": "Freightliner", "year": 2020,
            "vin": "1FUJGLD55LLAA1234", "plate_number": "ABC1234",
            "state_registered": "TX", "unit_number": "101",
        })

        tid = client.get(f"/api/trucks?company_id={company_id}").json()[0]["id"]

        d = rows(client, "/api/export/drivers", [did])
        assert d[0] == ["First name", "Middle name", "Last name", "DOB", "License state",
                        "License no", "Expire date", "Social", "Hire date", "Company"], d[0]
        row = d[1]
        assert row[:4] == ["Bob", "Q", "Rae", "01/03/1990"], row
        assert row[4] == "TX", row                       # kept, and upper-cased
        assert row[5] == "TX-777" and row[6] == "01/31/2031", row
        # Text, not a number — Excel must keep the leading zero of an SSN.
        assert row[7] == "012-34-5678", row
        assert row[8] == "03/04/2024" and row[9] == CO, row

        t = rows(client, "/api/export/trucks", [tid])
        assert t[0][0] == "Unit number" and t[1] == [
            "101", "Freightliner", "2020", "1FUJGLD55LLAA1234", "ABC1234", "TX", CO], t

        # No ids at all = the whole table; an empty list = the page is showing nothing.
        c = rows(client, "/api/export/companies")
        mine = [r for r in c if r[0] == CO][0]
        assert mine == [CO, "1234567", "MC-9999", "1 Main St, Dallas, TX 75001",
                        "(214) 555-0100", "Jane Owner", "02/03/1980"], mine
        assert len(rows(client, "/api/export/companies", [])) == 1, "headers only"
        assert len(rows(client, "/api/export/companies", [company_id])) == 2

        assert client.post("/api/export/nope", json={}).status_code == 404
        print("export OK")
    finally:
        app.dependency_overrides.clear()
        with Session(engine) as s:
            _clean(s)


if __name__ == "__main__":
    main()
